#!/usr/bin/env python3
"""generate-scorecard-xlsx.py

Generates the Website Auditor by Legion Code Inc. branded audit scorecard
workbook: skills/audit-scoring-stinger/references/templates/website-audit-scorecard-template.xlsx

This is a REAL, working openpyxl generator, not a spec document. Running it produces an
openable .xlsx with live formulas. Re-run any time this pair's design changes; never
hand-edit the generated .xlsx directly (per the template-injection discipline in
references/research/distilled-audit-scoring.md section 3 - the script, not a human editing
Excel, is the source of truth for this template's structure).

WHERE EACH DESIGN DECISION COMES FROM (see guides/03-nesting-structure-design.md for the
full writeup):

  - The 0-6 leaf scale, the 8 category weights in their exact descending order, the
    sub-category splits that ARE given numbers (Revenue Drivers 7/6/5, Analytics 5/4/3,
    Technical Deployment 3/4/4, Search Presence 3.5/3.5/2), the Content Score sub-audit
    NAMES (depth/freshness/coverage), the Security sub-audit NAMES (headers/TLS/cookies/
    CSP/platform exposure/client-side injection/payment-path integrity), the letter-grade
    threshold table, and the critical-security-override rule (Security leaf scores 1 ->
    final grade capped at C) are all directly SOURCED from
    plan/website-auditor-build-plan.md section 4 and prd-020-audit-scoring-index.md. Not
    invented here.

  - The masked, N/A-aware SUMPRODUCT formula shape (dual numerator/denominator mask,
    /SUM(weights) self-correcting divisor, zero-denominator IF guard) is SOURCED from
    skills/audit-scoring-stinger/references/research/distilled-audit-scoring.md sections 1-2
    (datacamp.com, exceljet.net, autorubric.org, in turn sourced from
    references/research/raw/).

  - The leaf -> sub-audit -> category -> final NESTING STRUCTURE itself (how many
    sub-audits per category where the build plan gives no number, how many example leaf
    rows per sub-audit, the "compact mirror" row trick used to make a scattered rollup
    column SUMPRODUCT-able against a compact named range, the column layout, the helper
    columns, the override-trigger lookup mechanism) is THIS SCRIPT'S OWN ENGINEERING
    DESIGN. The distilled research explicitly flags this nesting structure as an inference,
    not a directly sourced claim (distilled-audit-scoring.md section 4 and section 8). Every
    sub-audit name not given literally in the build plan (Mission Critical's three,
    Foundational Completeness's three) is an invented, illustrative placeholder, clearly
    labelled as such below, meant to be refined once a real engagement's leaf inventory
    exists.

  - Brand colors, the footer credit line, and the placeholder website URL are reused
    verbatim from the sibling audit-reporting-stinger's own placeholder brand config
    (skills/audit-reporting-stinger/references/templates/brand.json), itself explicitly
    marked as a placeholder pending the real Legion Code Inc. brand assets
    (brand/legion.css, brand/colors_and_type.css) that prd-020's open question notes were
    never vendored into this repo. Kept consistent across both Stingers rather than
    inventing a second, different placeholder palette.

Usage:
    python3 generate-scorecard-xlsx.py [output_path]

Default output path (when run with no args) is the template location this Stinger ships:
    ../templates/website-audit-scorecard-template.xlsx (relative to this script)
"""

import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName

# ---------------------------------------------------------------------------
# Brand constants (reused from skills/audit-reporting-stinger/references/templates/brand.json,
# itself an explicitly-flagged placeholder - see module docstring)
# ---------------------------------------------------------------------------

BRAND_NAME = "Legion Code Inc."
BRAND_PRODUCT = "Website Auditor"
BRAND_WEB = "https://legioncodeinc.com"
BRAND_CREDIT_LINE = "Audit tool created by Legion Code Inc."
BRAND_PRIMARY = "14213D"   # dark navy
BRAND_ACCENT = "2F6FED"    # blue
BRAND_MUTED = "6B7280"     # gray
BRAND_SURFACE = "F7F8FA"   # light gray-white

FILL_PRIMARY = PatternFill("solid", fgColor=BRAND_PRIMARY)
FILL_ACCENT = PatternFill("solid", fgColor=BRAND_ACCENT)
FILL_SURFACE = PatternFill("solid", fgColor=BRAND_SURFACE)
FILL_NA_GRAY = PatternFill("solid", fgColor="E5E7EB")
FILL_HEADER = PatternFill("solid", fgColor="DCE4F0")
FILL_WARN = PatternFill("solid", fgColor="FDE68A")
FILL_CRITICAL = PatternFill("solid", fgColor="FCA5A5")

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=BRAND_MUTED)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=BRAND_PRIMARY)
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=BRAND_PRIMARY)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_FOOTER = Font(name="Calibri", size=9, italic=True, color=BRAND_MUTED)
FONT_BIG_GRADE = Font(name="Calibri", size=36, bold=True, color=BRAND_PRIMARY)

THIN = Side(style="thin", color="D1D5DB")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GENERATED_AT = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

# ---------------------------------------------------------------------------
# Scoring design data
# SOURCED weights/names annotated inline; invented placeholder sub-audit names annotated
# "INVENTED" per the module docstring's grounding note.
# ---------------------------------------------------------------------------

LEAVES_PER_SUBAUDIT = 4  # illustrative leaf-row capacity per sub-audit in this template; engineering choice, see guides/03

# Category order is the EXACT descending order stated in the build plan / PRD-020.
CATEGORIES = [
    dict(
        key="security", name="Security", weight=20,
        subaudits=[  # names SOURCED (build plan section 4.2 "Contains" column); equal-split weights INVENTED (no split given)
            ("headers", "Headers", None),
            ("tls", "TLS", None),
            ("cookies", "Cookies", None),
            ("csp", "Content-Security-Policy", None),
            ("platform_exposure", "Platform Exposure", None),
            ("client_injection", "Client-Side Injection", None),
            ("payment_integrity", "Payment-Path Integrity", None),
        ],
    ),
    dict(
        key="revenue", name="Revenue Drivers", weight=18,
        subaudits=[  # names AND weights SOURCED (build plan section 4.2)
            ("visual_ux", "Visual UX/UI", 7),
            ("nav_journey", "Navigation & User Journey", 6),
            ("onpage_copy", "On-Page Copy", 5),
        ],
    ),
    dict(
        key="mission", name="Mission Critical", weight=14,
        subaudits=[  # INVENTED sub-audit names; PRD/build plan give only the gloss "does the site do the one job it exists to do"
            ("conversion_path", "Primary Conversion Path Integrity", None),
            ("core_reliability", "Core Function Reliability", None),
            ("trust_signals", "Trust & Credibility Signals", None),
        ],
    ),
    dict(
        key="analytics", name="Analytics and Insight", weight=12,
        subaudits=[  # names AND weights SOURCED (build plan section 4.2)
            ("foundational", "Foundational Analytics", 5),
            ("industry", "Industry-Specific Analytics", 4),
            ("deanonymization", "De-Anonymization (where lawful)", 3),
        ],
    ),
    dict(
        key="technical", name="Technical Deployment", weight=11,
        subaudits=[  # names AND weights SOURCED (build plan section 4.2)
            ("cdn", "CDN", 3),
            ("caching", "Caching Strategy", 4),
            ("cwv", "Core Web Vitals", 4),
        ],
    ),
    dict(
        key="foundational", name="Foundational Completeness", weight=10,
        subaudits=[  # INVENTED sub-audit names; PRD/build plan give only the gloss "the table stakes"
            ("core_pages", "Core Pages & Contact Info", None),
            ("legal_policy", "Legal & Policy Presence", None),
            ("basic_hygiene", "Basic Technical Hygiene", None),
        ],
    ),
    dict(
        key="search", name="Search Presence", weight=9,
        subaudits=[  # names AND weights SOURCED (build plan section 4.2)
            ("tech_seo", "Technical SEO", 3.5),
            ("tech_aeo", "Technical AEO", 3.5),
            ("subjective_copy", "Subjective Copy Read", 2),
        ],
    ),
    dict(
        key="content", name="Content Score", weight=6,
        subaudits=[  # names SOURCED (build plan section 4.2: "depth, freshness, coverage"); equal-split weights INVENTED
            ("depth", "Depth", None),
            ("freshness", "Freshness", None),
            ("coverage", "Coverage", None),
        ],
    ),
]

for _cat in CATEGORIES:
    _n = len(_cat["subaudits"])
    _cat["subaudits"] = [
        (k, n, w if w is not None else round(_cat["weight"] / _n, 4))
        for (k, n, w) in _cat["subaudits"]
    ]

# 0-6 leaf scale, SOURCED verbatim from build plan section 4.1
SCALE = [
    (0, "N/A", "no-op", "Audit point not relevant to this site type. Excluded from both numerator and denominator. Never counts as a failure."),
    (1, "F", "Critical", "Absent entirely, or present and critically failing. Blocks revenue, exposes risk, or breaks the function it exists to serve."),
    (2, "D", "High", "Present but materially broken. Works in some cases and fails in common ones."),
    (3, "C", "Medium", "Present and meets baseline. Low-severity findings only. Does the job without doing it well."),
    (4, "B minus", "Low", "Solid implementation. Minor findings a specialist would notice and a customer would not."),
    (5, "B", "Cosmetic", "Strong. Only cosmetic or preference-level findings remain."),
    (6, "A", "None", "Complete. Zero findings low through critical. Meets or exceeds the current published standard for this checkpoint."),
]

# Letter grade thresholds, SOURCED verbatim from build plan section 4.3 (ascending, ready for VLOOKUP approximate match)
GRADE_THRESHOLDS = [
    (0.00, "F"), (0.60, "D"), (0.70, "C-"), (0.73, "C"), (0.77, "C+"),
    (0.80, "B-"), (0.83, "B"), (0.87, "B+"), (0.90, "A-"), (0.93, "A"),
]

SECURITY_OVERRIDE_CAP = 0.7699  # just under the C+ threshold (0.77): a capped score can resolve to C or C-/D/F, never C+ or above.


def sanitize_name(*parts):
    """Excel defined-name-safe identifier."""
    s = "_".join(parts)
    s = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in s)
    if s and s[0].isdigit():
        s = "_" + s
    return s


def add_footer(ws, row, last_col_letter):
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = (
        f"{BRAND_CREDIT_LINE}  |  [LOGO PLACEHOLDER: Legion Code Inc. mark]  |  "
        f"{BRAND_WEB}  (placeholder URL, see brand.json note - real brand assets not yet vendored into this repo)"
    )
    cell.font = FONT_FOOTER
    cell.alignment = Alignment(horizontal="center")
    try:
        cell.hyperlink = BRAND_WEB
    except Exception:
        pass
    ws.row_dimensions[row].height = 18


def sheet_title_bar(ws, title, subtitle, last_col_letter, row=1):
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = FILL_PRIMARY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    if subtitle:
        ws.merge_cells(f"A{row+1}:{last_col_letter}{row+1}")
        c2 = ws.cell(row=row + 1, column=1, value=subtitle)
        c2.font = FONT_SUBTITLE
        c2.alignment = Alignment(horizontal="left", indent=1)
    return row + (3 if subtitle else 2)


# ===========================================================================
def build_workbook():
    wb = Workbook()

    named_ranges = {}  # name -> ref string, collected for the Config sheet listing

    def define_name(name, ref):
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names[name] = dn
        named_ranges[name] = ref

    # -----------------------------------------------------------------
    # 1. Cover
    # -----------------------------------------------------------------
    cover = wb.active
    cover.title = "Cover"
    cover.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", [4, 28, 40, 28, 4]):
        cover.column_dimensions[col].width = width
    row = sheet_title_bar(cover, f"{BRAND_PRODUCT} - Audit Scorecard", "Prepared by " + BRAND_NAME, "E", row=3)
    fields = [
        ("Auditor", "{{auditor_name}}"),
        ("Audited party contact", "{{contact_name}}"),
        ("Audited party business", "{{business_name}}"),
        ("Domain", "{{website_url}}"),
        ("Engagement reference", "{{engagement_ref}}"),
        ("Report date", "{{report_date}}"),
        ("Template generated", GENERATED_AT),
        ("Template generator", "generate-scorecard-xlsx.py (audit-scoring-stinger)"),
    ]
    for i, (label, val) in enumerate(fields):
        r = row + i
        cover.cell(row=r, column=2, value=label).font = FONT_BOLD
        cover.cell(row=r, column=3, value=val).font = FONT_BODY
    note_row = row + len(fields) + 2
    cover.merge_cells(f"B{note_row}:D{note_row + 3}")
    cover.cell(row=note_row, column=2, value=(
        "{{...}} placeholders are hydrated by audit-intake-worker-bee's four recorded answers "
        "and by audit-scoring-worker-bee at rollup time. This Cover sheet carries no formulas; "
        "it is populated by direct cell write, per the template-injection pattern in "
        "references/research/distilled-audit-scoring.md section 3."
    )).font = FONT_SUBTITLE
    cover.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    add_footer(cover, note_row + 6, "E")

    # -----------------------------------------------------------------
    # 2. Rubric  (the named-range source of truth - AC-4)
    # -----------------------------------------------------------------
    rub = wb.create_sheet("Rubric")
    rub.sheet_view.showGridLines = False
    widths = {"A": 26, "B": 34, "C": 14, "D": 60, "E": 14, "F": 30}
    for col, w in widths.items():
        rub.column_dimensions[col].width = w
    r = sheet_title_bar(rub, "Rubric - Scoring Configuration", "Every weight used by every formula in this workbook lives here as a named range. Retune here; never edit a formula.", "F", row=1)

    # 0-6 scale table
    rub.cell(row=r, column=1, value="0-6 Leaf Scale").font = FONT_SECTION
    r += 1
    headers = ["Value", "Grade", "Band", "Definition"]
    for i, h in enumerate(headers):
        c = rub.cell(row=r, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    r += 1
    scale_first_row = r
    for value, grade, band, definition in SCALE:
        rub.cell(row=r, column=1, value=value).border = BORDER_ALL
        rub.cell(row=r, column=2, value=grade).border = BORDER_ALL
        rub.cell(row=r, column=3, value=band).border = BORDER_ALL
        cdef = rub.cell(row=r, column=4, value=definition)
        cdef.border = BORDER_ALL
        cdef.alignment = Alignment(wrap_text=True)
        r += 1
    scale_last_row = r - 1
    r += 2

    # Category weight table (the CategoryWeights named range - the single most load-bearing range in the workbook)
    rub.cell(row=r, column=1, value="Category Weights (points, sums to 100)").font = FONT_SECTION
    r += 1
    for i, h in enumerate(["Rank", "Category", "Weight", "Named range", "Check"]):
        c = rub.cell(row=r, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    r += 1
    cat_weight_first_row = r
    cat_wt_named_names = []
    for idx, cat in enumerate(CATEGORIES, start=1):
        rub.cell(row=r, column=1, value=idx).border = BORDER_ALL
        rub.cell(row=r, column=2, value=cat["name"]).border = BORDER_ALL
        wcell = rub.cell(row=r, column=3, value=cat["weight"])
        wcell.border = BORDER_ALL
        wcell.font = FONT_BOLD
        nm = sanitize_name("Wt", cat["key"])
        define_name(nm, f"'Rubric'!${get_column_letter(3)}${r}")
        rub.cell(row=r, column=4, value=nm).font = FONT_SUBTITLE
        cat["_weight_cell"] = f"'Rubric'!$C${r}"
        cat["_weight_row"] = r
        cat_wt_named_names.append(nm)
        r += 1
    cat_weight_last_row = r - 1
    define_name("CategoryWeights", f"'Rubric'!$C${cat_weight_first_row}:$C${cat_weight_last_row}")
    check_row = r
    rub.cell(row=r, column=2, value="Check: should equal 100").font = FONT_SUBTITLE
    ccell = rub.cell(row=r, column=3, value=f"=SUM(C{cat_weight_first_row}:C{cat_weight_last_row})")
    ccell.font = FONT_BOLD
    rub.conditional_formatting.add(
        f"C{check_row}",
        FormulaRule(formula=[f"C{check_row}<>100"], fill=FILL_CRITICAL),
    )
    r += 3

    # Sub-audit weight tables, one block per category
    rub.cell(row=r, column=1, value="Sub-audit Weights (per category; self-normalizing, do not need to sum to the category weight)").font = FONT_SECTION
    r += 1
    for cat in CATEGORIES:
        rub.cell(row=r, column=1, value=cat["name"]).font = FONT_BOLD
        rub.cell(row=r, column=1).fill = FILL_SURFACE
        r += 1
        sub_first_row = r
        for key, name, weight in cat["subaudits"]:
            rub.cell(row=r, column=2, value=name).border = BORDER_ALL
            wc = rub.cell(row=r, column=3, value=weight)
            wc.border = BORDER_ALL
            r += 1
        sub_last_row = r - 1
        subwt_name = sanitize_name("SubWt", cat["key"])
        define_name(subwt_name, f"'Rubric'!$C${sub_first_row}:$C${sub_last_row}")
        rub.cell(row=r, column=2, value=f"Named range: {subwt_name}").font = FONT_SUBTITLE
        cat["_subwt_name"] = subwt_name
        r += 2

    # Letter grade threshold table
    rub.cell(row=r, column=1, value="Letter Grade Thresholds").font = FONT_SECTION
    r += 1
    for i, h in enumerate(["Min %", "Letter"]):
        c = rub.cell(row=r, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    r += 1
    grade_first_row = r
    for minpct, letter in GRADE_THRESHOLDS:
        c1 = rub.cell(row=r, column=1, value=minpct)
        c1.number_format = "0%"
        c1.border = BORDER_ALL
        rub.cell(row=r, column=2, value=letter).border = BORDER_ALL
        r += 1
    grade_last_row = r - 1
    define_name("GradeTable", f"'Rubric'!$A${grade_first_row}:$B${grade_last_row}")
    r += 2

    # Critical security override configuration
    rub.cell(row=r, column=1, value="Critical Security Override").font = FONT_SECTION
    r += 1
    rub.cell(row=r, column=1, value="Override enabled").font = FONT_BODY
    ecell = rub.cell(row=r, column=2, value=True)
    define_name("SecurityOverrideEnabled", f"'Rubric'!$B${r}")
    r += 1
    rub.cell(row=r, column=1, value="Cap (max % when override active)").font = FONT_BODY
    capcell = rub.cell(row=r, column=2, value=SECURITY_OVERRIDE_CAP)
    capcell.number_format = "0.00%"
    define_name("SecurityOverrideCapPct", f"'Rubric'!$B${r}")
    r += 2
    rub.merge_cells(f"A{r}:F{r+2}")
    rub.cell(row=r, column=1, value=(
        "Any Security-category leaf scored 1 sets the override flag on Executive Scorecard, which "
        "caps the FINAL grade at the letter this Cap % resolves to via GradeTable (default just under "
        "the C+ threshold, i.e. C). This is a ceiling (MIN), never a floor: a final score that is "
        "already below the cap is left alone. Source: build plan section 4.3 Q9, adopted as-is per prd-020."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    rub.cell(row=r, column=1).font = FONT_SUBTITLE
    r += 4

    footer_row_rubric = r + 1
    add_footer(rub, footer_row_rubric, "F")

    # -----------------------------------------------------------------
    # 3. Scorecard (the rollup engine: leaf -> sub-audit -> category)
    # -----------------------------------------------------------------
    sc = wb.create_sheet("Scorecard")
    sc.sheet_view.showGridLines = False
    sc.freeze_panes = "A4"
    col_widths = {"A": 18, "B": 26, "C": 12, "D": 42, "E": 9, "F": 26, "G": 30, "H": 9, "I": 16, "J": 18, "K": 3, "L": 11, "M": 6}
    for col, w in col_widths.items():
        sc.column_dimensions[col].width = w

    r = sheet_title_bar(
        sc,
        "Scorecard - Leaf to Sub-audit to Category Rollup",
        "N/A-aware masked SUMPRODUCT at every level. Column L carries every rollup value; column E carries only real leaf scores.",
        "M", row=1,
    )
    header_row = r
    headers = ["Category", "Sub-audit", "Leaf ID", "Description", "Score (0-6)", "Evidence pointer",
               "Justification", "Weight", "Included?", "Row type", "", "Rollup %", ""]
    for i, h in enumerate(headers):
        c = sc.cell(row=header_row, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    sc.freeze_panes = f"A{header_row+1}"
    r = header_row + 1

    category_rollup_row = {}
    category_block_bounds = {}   # key -> (start_row, end_row) inclusive, whole category block
    example_cycle = [6, 6, 6, 0]  # three strong scores plus one deliberate N/A per sub-audit: makes the N/A-mask math easy to hand-verify (3*6/(3*6)=100% when masked correctly, vs the wrong 3*6/(4*6)=75% if N/A were not excluded), and lets the critical-security-override demo below show a dramatic before/after (would-be A, capped to C) rather than a marginal one

    leaf_counter = 0
    for cat in CATEGORIES:
        cat_start = r
        sc.cell(row=r, column=1, value=cat["name"]).font = FONT_BOLD
        sc.cell(row=r, column=1).fill = FILL_SURFACE
        for col in range(2, 14):
            sc.cell(row=r, column=col).fill = FILL_SURFACE
        sc.cell(row=r, column=3, value=f"Weight: {cat['weight']} pts  (={cat['_weight_cell']})").font = FONT_SUBTITLE
        sc.cell(row=r, column=10, value="Category Header")
        r += 1

        subaudit_rollup_rows = []
        for sub_key, sub_name, sub_weight in cat["subaudits"]:
            sc.cell(row=r, column=2, value=f"{sub_name} (weight {sub_weight})").font = FONT_BOLD
            sc.cell(row=r, column=10, value="Sub-audit Header")
            r += 1
            leaf_start = r
            for i in range(LEAVES_PER_SUBAUDIT):
                leaf_counter += 1
                score = example_cycle[i % len(example_cycle)]
                # Deliberately trigger the critical-security-override demo on the first Security sub-audit's first leaf.
                if cat["key"] == "security" and sub_key == "headers" and i == 0:
                    score = 1
                    desc = "Example: Strict-Transport-Security header absent (illustrative - demonstrates the override firing)"
                elif score == 0:
                    desc = f"Example leaf {i+1} for {sub_name} (illustrative - not applicable to this site type)"
                else:
                    desc = f"Example leaf {i+1} for {sub_name} (illustrative placeholder - replaced by the real upstream Bee's finding at run time)"
                sc.cell(row=r, column=1, value=cat["name"])
                sc.cell(row=r, column=2, value=sub_name)
                leaf_id = f"{cat['key'].upper()}-{sub_key.upper()}-{i+1:02d}"
                sc.cell(row=r, column=3, value=leaf_id)
                sc.cell(row=r, column=4, value=desc).alignment = Alignment(wrap_text=True)
                scell = sc.cell(row=r, column=5, value=score)
                scell.alignment = Alignment(horizontal="center")
                sc.cell(row=r, column=6, value="example-evidence-pointer.md#L1" if score else "n/a")
                sc.cell(row=r, column=7, value="Illustrative justification text." if score else "Not applicable to this site type.")
                wcell = sc.cell(row=r, column=8, value=1)
                wcell.alignment = Alignment(horizontal="center")
                inc = sc.cell(row=r, column=9, value=f'=IF(E{r}="","",IF(E{r}>0,"Included","Excluded (N/A)"))')
                sc.cell(row=r, column=10, value="Leaf")
                trig = sc.cell(row=r, column=13, value=f'=IF(E{r}=1,1,0)')
                for col in range(1, 14):
                    sc.cell(row=r, column=col).border = BORDER_ALL
                r += 1
            leaf_end = r - 1

            rollup_formula = (
                f'=IF(SUMPRODUCT(--(E{leaf_start}:E{leaf_end}>0),H{leaf_start}:H{leaf_end})=0,"N/A",'
                f'SUMPRODUCT(--(E{leaf_start}:E{leaf_end}>0),E{leaf_start}:E{leaf_end},H{leaf_start}:H{leaf_end})'
                f'/(6*SUMPRODUCT(--(E{leaf_start}:E{leaf_end}>0),H{leaf_start}:H{leaf_end})))'
            )
            sc.cell(row=r, column=2, value=f"  -> {sub_name} rollup").font = FONT_SUBTITLE
            rcell = sc.cell(row=r, column=12, value=rollup_formula)
            rcell.font = FONT_BOLD
            rcell.number_format = "0.0%"
            sc.cell(row=r, column=10, value="Sub-audit Rollup")
            subaudit_rollup_rows.append(r)
            r += 1

        # compact mirror table (contiguous, aligns 1:1 with the Rubric SubWt_<key> named range)
        mirror_start = r
        for (sub_key, sub_name, sub_weight), src_row in zip(cat["subaudits"], subaudit_rollup_rows):
            sc.cell(row=r, column=2, value=f"[mirror] {sub_name}").font = FONT_SUBTITLE
            mcell = sc.cell(row=r, column=12, value=f"=L{src_row}")
            mcell.number_format = "0.0%"
            sc.cell(row=r, column=10, value="Sub-audit Mirror")
            r += 1
        mirror_end = r - 1

        cat_rollup_formula = (
            f"=IF(SUM({cat['_subwt_name']})=0,\"N/A\","
            f"SUMPRODUCT(L{mirror_start}:L{mirror_end},{cat['_subwt_name']})/SUM({cat['_subwt_name']}))"
        )
        sc.cell(row=r, column=1, value=f"{cat['name']} - CATEGORY ROLLUP").font = FONT_BOLD
        crcell = sc.cell(row=r, column=12, value=cat_rollup_formula)
        crcell.font = FONT_BOLD
        crcell.number_format = "0.0%"
        crcell.fill = FILL_HEADER
        sc.cell(row=r, column=10, value="Category Rollup")
        category_rollup_row[cat["key"]] = r
        cat_end = r
        category_block_bounds[cat["key"]] = (cat_start, cat_end)
        r += 2

    last_data_row = r
    # Conditional formatting on leaf score column
    score_range = f"E{header_row+1}:E{last_data_row}"
    sc.conditional_formatting.add(
        score_range,
        FormulaRule(formula=[f"AND(E{header_row+1}=0,E{header_row+1}<>\"\")"], fill=FILL_NA_GRAY),
    )
    sc.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="num", start_value=1, start_color="F8696B",
            mid_type="num", mid_value=3.5, mid_color="FFEB84",
            end_type="num", end_value=6, end_color="63BE7B",
        ),
    )
    rollup_range = f"L{header_row+1}:L{last_data_row}"
    sc.conditional_formatting.add(
        rollup_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=0.5, mid_color="FFEB84",
            end_type="num", end_value=1, end_color="63BE7B",
        ),
    )

    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=6, allow_blank=True,
                         error="Score must be a whole number 0-6.", errorTitle="Invalid score")
    sc.add_data_validation(dv)
    dv.add(score_range)

    add_footer(sc, last_data_row + 2, "M")

    # -----------------------------------------------------------------
    # 4. Audit Tree (pointer sheet - see guides/03 for why this is consolidated into Scorecard)
    # -----------------------------------------------------------------
    tree = wb.create_sheet("Audit Tree")
    tree.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [4, 90, 4, 4, 4]):
        tree.column_dimensions[col].width = w
    r = sheet_title_bar(tree, "Audit Tree", None, "E", row=2)
    tree.merge_cells(f"B{r}:D{r+6}")
    tree.cell(row=r, column=2, value=(
        "This template consolidates the full leaf -> sub-audit -> category hierarchy into the "
        "'Scorecard' sheet rather than duplicating it on a second sheet, so there is exactly one "
        "place the rollup formulas live and one place to audit them. This sheet name exists to "
        "satisfy build plan section 4.4's sheet inventory; its content is the design decision, "
        "documented in guides/03-nesting-structure-design.md, not an omission. Go to 'Scorecard' "
        "for the full hierarchy, every leaf, its score, weight, evidence pointer, and justification."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    add_footer(tree, r + 8, "E")

    # -----------------------------------------------------------------
    # 5. One linked view sheet per category (Security ... Content Score)
    # -----------------------------------------------------------------
    for cat in CATEGORIES:
        cs_start, cs_end = category_block_bounds[cat["key"]]
        ws = wb.create_sheet(cat["name"][:31])
        ws.sheet_view.showGridLines = False
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
        r = sheet_title_bar(ws, f"{cat['name']} - Detail View", f"Live-linked view of Scorecard rows {cs_start}-{cs_end}. Edit scores on 'Scorecard'; this sheet mirrors, it does not recompute.", "M", row=1)
        for i, h in enumerate(headers):
            c = ws.cell(row=r, column=1 + i, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
        r += 1
        for src_row in range(cs_start, cs_end + 1):
            for col in range(1, 13):
                col_letter = get_column_letter(col)
                ws.cell(row=r, column=col, value=f"=Scorecard!{col_letter}{src_row}")
            r += 1
        add_footer(ws, r + 1, "M")

    # -----------------------------------------------------------------
    # 6. Executive Scorecard (final rollup, letter grade, override banner)
    # -----------------------------------------------------------------
    ex = wb.create_sheet("Executive Scorecard", 1)  # index 1: right after Cover
    ex.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [26, 14, 14, 18, 40]):
        ex.column_dimensions[col].width = w
    r = sheet_title_bar(ex, "Executive Scorecard", "Final grade, per-category rollups, and the critical-security-override banner.", "E", row=1)

    ex.cell(row=r, column=1, value="Category").font = FONT_HEADER
    ex.cell(row=r, column=2, value="Weight").font = FONT_HEADER
    ex.cell(row=r, column=3, value="Rollup %").font = FONT_HEADER
    ex.cell(row=r, column=4, value="Contribution").font = FONT_HEADER
    for c in range(1, 5):
        ex.cell(row=r, column=c).fill = FILL_HEADER
    r += 1
    summary_first_row = r
    for cat in CATEGORIES:
        ex.cell(row=r, column=1, value=cat["name"])
        wcell = ex.cell(row=r, column=2, value=f"={sanitize_name('Wt', cat['key'])}")
        rcell = ex.cell(row=r, column=3, value=f"=Scorecard!L{category_rollup_row[cat['key']]}")
        rcell.number_format = "0.0%"
        ccell = ex.cell(row=r, column=4, value=f"=B{r}*C{r}")
        ccell.number_format = "0.00"
        for col in range(1, 5):
            ex.cell(row=r, column=col).border = BORDER_ALL
        r += 1
    summary_last_row = r - 1

    final_row = r + 1
    ex.cell(row=final_row, column=1, value="FINAL SCORE (raw, before override)").font = FONT_BOLD
    final_cell = ex.cell(
        row=final_row, column=3,
        value=f"=SUMPRODUCT(C{summary_first_row}:C{summary_last_row},CategoryWeights)/SUM(CategoryWeights)",
    )
    final_cell.number_format = "0.0%"
    final_cell.font = FONT_BOLD

    # Critical security override block: scan the Security category's own block on Scorecard
    sec_start, sec_end = category_block_bounds["security"]
    override_row = final_row + 2
    ex.cell(row=override_row, column=1, value="Critical security override active?").font = FONT_BOLD
    override_cell = ex.cell(
        row=override_row, column=3,
        value=f"=IF(AND(SecurityOverrideEnabled,COUNTIFS(Scorecard!$E${sec_start}:$E${sec_end},1)>0),TRUE,FALSE)",
    )

    trigger_row = override_row + 1
    ex.cell(row=trigger_row, column=1, value="Triggering finding (leaf ID)").font = FONT_BODY
    trig_id_cell = ex.cell(
        row=trigger_row, column=3,
        value=(
            f'=IFERROR(INDEX(Scorecard!$C${sec_start}:$C${sec_end},'
            f'MATCH(1,Scorecard!$M${sec_start}:$M${sec_end},0)),"")'
        ),
    )
    ex.cell(row=trigger_row + 1, column=1, value="Triggering finding (description)").font = FONT_BODY
    trig_desc_cell = ex.cell(
        row=trigger_row + 1, column=3,
        value=(
            f'=IFERROR(INDEX(Scorecard!$D${sec_start}:$D${sec_end},'
            f'MATCH(1,Scorecard!$M${sec_start}:$M${sec_end},0)),"")'
        ),
    )

    effective_row = trigger_row + 3
    ex.cell(row=effective_row, column=1, value="Effective % (capped if override active)").font = FONT_BOLD
    effective_cell = ex.cell(
        row=effective_row, column=3,
        value=f"=IF(C{override_row},MIN(C{final_row},SecurityOverrideCapPct),C{final_row})",
    )
    effective_cell.number_format = "0.0%"
    effective_cell.font = FONT_BOLD

    grade_row = effective_row + 2
    ex.cell(row=grade_row, column=1, value="FINAL LETTER GRADE").font = FONT_SECTION
    grade_cell = ex.cell(row=grade_row, column=3, value=f"=VLOOKUP(C{effective_row},GradeTable,2,TRUE)")
    grade_cell.font = FONT_BIG_GRADE
    ex.row_dimensions[grade_row].height = 42

    banner_row = grade_row + 2
    ex.merge_cells(f"A{banner_row}:E{banner_row+2}")
    banner_cell = ex.cell(
        row=banner_row, column=1,
        value=(
            f'=IF(C{override_row},"CRITICAL SECURITY OVERRIDE ACTIVE - final grade capped at "'
            f'&VLOOKUP(SecurityOverrideCapPct,GradeTable,2,TRUE)&". Triggering finding: "'
            f'&C{trigger_row}&" - "&C{trigger_row+1},"No critical security override active.")'
        ),
    )
    banner_cell.alignment = Alignment(wrap_text=True, vertical="top")
    banner_cell.font = FONT_BOLD

    ex.conditional_formatting.add(f"C{override_row}", FormulaRule(formula=[f"C{override_row}=TRUE"], fill=FILL_CRITICAL))
    ex.conditional_formatting.add(f"A{banner_row}:E{banner_row+2}", FormulaRule(formula=[f"C{override_row}=TRUE"], fill=FILL_CRITICAL))
    ex.conditional_formatting.add(
        f"C{grade_row}",
        FormulaRule(formula=[f'OR(C{grade_row}="F",C{grade_row}="D")'], fill=FILL_CRITICAL),
    )
    ex.conditional_formatting.add(
        f"C{grade_row}",
        FormulaRule(formula=[f'LEFT(C{grade_row},1)="C"'], fill=FILL_WARN),
    )

    add_footer(ex, banner_row + 4, "E")

    # -----------------------------------------------------------------
    # 7. Findings Register
    # -----------------------------------------------------------------
    fr = wb.create_sheet("Findings Register")
    fr.sheet_view.showGridLines = False
    fr_headers = ["Finding ID", "Severity", "Category", "Sub-audit", "Page/URL", "Evidence", "Remediation", "Effort (hrs band)", "Status"]
    for col, w in zip("ABCDEFGHI", [14, 10, 18, 24, 30, 30, 40, 14, 12]):
        fr.column_dimensions[col].width = w
    r = sheet_title_bar(fr, "Findings Register", "Every finding, one row each. Populated by audit-scoring-worker-bee from upstream Bees' leaf findings.", "I", row=1)
    for i, h in enumerate(fr_headers):
        c = fr.cell(row=r, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    r += 1
    example = ["SEC-HEADERS-01", "Critical", "Security", "Headers", "https://example.com/",
               "example-evidence-pointer.md#L1", "Add Strict-Transport-Security header.", "1-4", "Open"]
    for i, v in enumerate(example):
        fr.cell(row=r, column=1 + i, value=v).border = BORDER_ALL
    r += 2
    add_footer(fr, r + 1, "I")

    # -----------------------------------------------------------------
    # 8. Evidence Index
    # -----------------------------------------------------------------
    ei = wb.create_sheet("Evidence Index")
    ei.sheet_view.showGridLines = False
    ei_headers = ["Evidence ID", "Artifact path", "Produced by", "Captured at", "Description"]
    for col, w in zip("ABCDE", [14, 40, 26, 18, 40]):
        ei.column_dimensions[col].width = w
    r = sheet_title_bar(ei, "Evidence Index", "Artifact map back to the shared audit workspace (build plan section 3).", "E", row=1)
    for i, h in enumerate(ei_headers):
        c = ei.cell(row=r, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    r += 2
    add_footer(ei, r + 1, "E")

    # -----------------------------------------------------------------
    # 9. Config
    # -----------------------------------------------------------------
    cfg = wb.create_sheet("Config")
    cfg.sheet_view.showGridLines = False
    for col, w in zip("ABC", [30, 50, 10]):
        cfg.column_dimensions[col].width = w
    r = sheet_title_bar(cfg, "Config", "Named ranges, version stamp, and generation metadata.", "C", row=1)
    cfg.cell(row=r, column=1, value="Template version").font = FONT_BOLD
    cfg.cell(row=r, column=2, value="1.0.0")
    r += 1
    cfg.cell(row=r, column=1, value="Generated at").font = FONT_BOLD
    cfg.cell(row=r, column=2, value=GENERATED_AT)
    r += 1
    cfg.cell(row=r, column=1, value="Generator script").font = FONT_BOLD
    cfg.cell(row=r, column=2, value="skills/audit-scoring-stinger/references/scripts/generate-scorecard-xlsx.py")
    r += 2
    cfg.cell(row=r, column=1, value="Named ranges").font = FONT_SECTION
    r += 1
    for i, h in enumerate(["Name", "Refers to"]):
        c = cfg.cell(row=r, column=1 + i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
    r += 1
    for name, ref in sorted(named_ranges.items()):
        cfg.cell(row=r, column=1, value=name).border = BORDER_ALL
        cfg.cell(row=r, column=2, value=ref).border = BORDER_ALL
        r += 1
    r += 1
    add_footer(cfg, r + 1, "C")

    wb.active = 0
    return wb


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "templates", "website-audit-scorecard-template.xlsx"
    )
    out_path = os.path.abspath(out_path)
    wb = build_workbook()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
